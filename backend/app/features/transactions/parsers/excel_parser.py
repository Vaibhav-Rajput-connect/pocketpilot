from __future__ import annotations

import io

from openpyxl import load_workbook

from app.features.transactions.parsers.base import (
    BaseParser,
    RawTransaction,
    detect_columns,
    rows_to_transactions,
)


class ExcelParser(BaseParser):
    """Parses Excel (.xlsx) bank statement files."""

    def parse(self, file_bytes: bytes, filename: str) -> list[RawTransaction]:
        wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active

        if ws is None:
            return []

        all_rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            all_rows.append([str(cell) if cell is not None else "" for cell in row])

        wb.close()

        if len(all_rows) < 2:
            return []

        # Find header row by looking for date and amount columns
        header_idx = -1
        col_map = None
        for i, row in enumerate(all_rows):
            temp_map = detect_columns(row)
            if temp_map["date"] is not None and (temp_map["amount"] is not None or temp_map["debit"] is not None):
                header_idx = i
                col_map = temp_map
                break

        if header_idx == -1 or col_map is None:
            return []

        data_rows = all_rows[header_idx + 1:]
        return rows_to_transactions(data_rows, col_map, filename)
